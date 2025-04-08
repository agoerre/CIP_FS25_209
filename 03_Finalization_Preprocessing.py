### This code is for finalizing the dataset for analysis by renaming columns, categorizing altitude, and saving the updated DataFrame to an Excel file. ###

# --- Step 0: Import necessary libraries ---
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# --- Step 1: Load the dataset and make a copy ---
# Replace with your actual file path
original_df = pd.read_csv("data_with_categories.csv")

# Create a working copy
df = original_df.copy()

# --- Step 2: Rename specific columns ---
df.rename(columns={
    "Type": "Location Type",
    "KÃ¶ppen climate type": "Climate Type",
    "Category": "Mineral Category"
}, inplace=True)

# --- Step 3: Define function to categorize altitude ---
def categorize_altitude(altitude):
    if altitude <= 600:
        return "Tiefland"
    elif 601 <= altitude <= 1000:
        return "Hügelland"
    elif 1001 <= altitude <= 1500:
        return "Mittelgebirge"
    elif 1501 <= altitude <= 2500:
        return "Hochgebirge"
    elif 2501 <= altitude <= 3000:
        return "Subnivale Zone"
    else:
        return "Nivale Zone"

# --- Step 4: Create the "Altitude Category" column ---
df["Altitude Category"] = df["Altitude"].apply(categorize_altitude)

# --- Step 5: Move the new column directly after "Altitude" ---
altitude_index = df.columns.get_loc("Altitude")
columns = list(df.columns)
# Remove the "Altitude Category" column and re-insert it right after "Altitude"
columns.insert(altitude_index + 1, columns.pop(columns.index("Altitude Category")))
df = df[columns]

# --- Step 6: Save the updated DataFrame to a new Excel file ---
excel_path = "03_Finalized_Dataset.xlsx"
df.to_excel(excel_path, index=False)

# --- Step 7: Auto-adjust column widths using openpyxl ---
wb = load_workbook(excel_path)
ws = wb.active

for col_idx, column_cells in enumerate(ws.columns, 1):
    max_length = 0
    column = get_column_letter(col_idx)
    for cell in column_cells:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2  # Add padding
    ws.column_dimensions[column].width = adjusted_width

wb.save(excel_path)
print("Finalized dataset saved as '03_Finalized_Dataset.xlsx'.")