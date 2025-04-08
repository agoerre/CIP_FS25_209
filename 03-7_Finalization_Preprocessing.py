### This code is for finalizing the dataset for analysis by renaming columns, categorizing altitude, and saving the updated DataFrame to an Excel file. ###

# --- STEP 0: IMPORT NECESSARY LIBRARIES ---
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# --- STEP 1: LOAD THE DATASET AND MAKE A COPY ---
original_df = pd.read_csv("Data/03-6_Minerals-Categorization_Cleaned-dataset.csv")

# Create a working copy
df = original_df.copy()


# --- STEP 2: RENAME SPECIFIC COLUMNS ---
df.rename(columns={
    "Type": "Location Type",
    "KÃ¶ppen climate type": "Climate Type",
    "Category": "Mineral Category"
}, inplace=True)


# --- STEP 3: DEFINE FUNCTION TO CATEGORIZE ALTITUDE ---
def categorize_altitude(altitude):
    if altitude <= 600:
        return "Tiefland"
    elif 601 <= altitude <= 1000:
        return "Hügelland"
    elif 1001 <= altitude <= 1500:
        return "Mittelgebirge"
    elif 1501 <= altitude <= 2500:
        return "Hochgebirge"
    elif 2501 <= altitude <= 3000:
        return "Subnivale Zone"
    else:
        return "Nivale Zone"


# --- STEP 4: CREATE THE "Altitude Category" COLUMN ---
df["Altitude Category"] = df["Altitude"].apply(categorize_altitude)


# --- STEP 5: MOVE THE NEW COLUMN DIRECTLY AFTER "Altitude" ---
altitude_index = df.columns.get_loc("Altitude")
columns = list(df.columns)
# Remove the "Altitude Category" column and re-insert it right after "Altitude"
columns.insert(altitude_index + 1, columns.pop(columns.index("Altitude Category")))
df = df[columns]


# --- STEP 6: SAVE THE UPDATED DATAFRAME TO A NEW EXCEL FILE ---
excel_path = "Data/03-8_Finalized_Dataset.xlsx"
df.to_excel(excel_path, index=False)


# --- STEP 7: AUTO-ADJUST COLUMN WIDTHS USING OPENPYXL ---
wb = load_workbook(excel_path)
ws = wb.active

for col_idx, column_cells in enumerate(ws.columns, 1):
    max_length = 0
    column = get_column_letter(col_idx)
    for cell in column_cells:
        try:
            max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    adjusted_width = max_length + 2  # Add padding
    ws.column_dimensions[column].width = adjusted_width

wb.save(excel_path)
print("Finalized dataset saved as '03-8_Finalized_Dataset.xlsx' in the Data subfolder.")